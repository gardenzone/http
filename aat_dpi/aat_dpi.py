# coding=gbk

import openpyxl
from yaml import load, FullLoader
import os
import sys
import time
from json import JSONDecodeError
import urllib3
import requests


# 配置参数
# 每个 case 最多的处理的 excel 数据量
max_modify_step = 10


# 读取 excel 文件，返回筛选后的数据
def read_dpi_config(file_path, data_select_num):
    default_file = file_path
    default_sheet_name = '本次变更'
    title = {
        'flag': '变更标志',
        'description': '变更说明',
        'code': '业务编码',
        'name': '业务命名',
        'l3': '三层目的IP地址',
        'protocol': '协议号',
        'l4': '四层目的端口号',
        'l7': '七层URL',
    }
    select_num = data_select_num

    file_path = default_file
    sheet_name = default_sheet_name
    # 打开 excel 文件
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    row_count = sheet.max_row
    # col_count = sheet.max_column

    def get_one_row(sheet, row_num, discard_none=False):
        col_count = sheet.max_column
        if discard_none:
            result = [sheet.cell(row_num, i).value for i in range(1, col_count + 1) if
                      sheet.cell(row_num, i).value != None]
        else:
            result = [sheet.cell(row_num, i).value for i in range(1, col_count + 1)]
        return result

    def init_title_col(title_list):
        for key, value in title.items():
            if value in title_list:
                title[key] = title_list.index(value) + 1

    # 读取表头数据，初始化列表头数据的列下标
    title_values1 = get_one_row(sheet, 1)
    title_values2 = get_one_row(sheet, 2)
    init_title_col(title_values1)
    init_title_col(title_values2)

    # print(title)

    # 对每行数据进行分类
    def rule_type(title_values, row_num):
        # layer = ['l3', 'l4', 'l7', 'protocol']
        layer = ['l3', 'l7']
        key_list = []

        # l3, l4, l7, protocol
        for key in layer:
            value = title_values.get(key)
            if value != None and value != "":
                key_list.append(key)
        # http/https
        if title_values.get('l7'):
            if 'https://' in title_values.get('l7'):
                key_list.append('https')
            else:
                key_list.append('http')

        type_str = "+".join(key_list)
        return type_str

        # 对每一行数据进行分类

    rule_types = {}
    for row_num in range(3, row_count):
        row_value = {'line_num': row_num}
        for name, col in title.items():
            row_value[name] = sheet.cell(row_num, col).value
        # 只选择变更标志为“新增”的规则，“删除”的规则不需要测试
        if row_value['flag'] != '新增':
            continue
        type_str = rule_type(row_value, row_num)

        code = row_value['code']
        if code not in rule_types:
            rule_types[code] = {}
        if type_str not in rule_types[code]:
            rule_types[code][type_str] = []
        rule_types[code][type_str].append(row_value)

    def mask_sort(line):
        ip_mask = line['l3']
        # 优先选唯一地址，其次选 ip 范围最小的
        if "/" not in ip_mask:
            return 0
        mask = int(ip_mask.split('/')[1])
        # ipv6
        if mask > 32:
            return 128 - mask
        # ipv4
        else:
            return 32 - mask

    def url_sort(line):
        # 优先选 * 号少的，其次选 * 在后面的
        url = line['l7']
        star_num = url.count("*")
        if star_num == 0:
            return 0
        elif star_num == 1:
            if url[-1] == "*":
                return 2
            else:
                return 1
        else:
            return 3

    # 根据设计的优先级对每类数据进行排序，抽取优先级最高的几条数据
    def case_choice(type_str, type_list):
        if len(type_list) < select_num:
            return type_str, type_list
        sort_list = []
        # 排序
        if 'l3' in type_str and 'l7' in type_str:
            sort_list = sorted(type_list, key=lambda line: (mask_sort(line), url_sort((line))))
        elif 'l3' in type_str:
            sort_list = sorted(type_list, key=mask_sort)
        elif 'l7' in type_str:
            sort_list = sorted(type_list, key=url_sort)
        # 抽取部分数据返回
        result = sort_list[:select_num]
        # result = sort_list
        return type_str, result

    # 将不同业务编码，但数据类型相同的数据合并到一起
    case_select = {}
    for code, types in rule_types.items():
        for type_str, tpye_list in types.items():
            t_str, selects = case_choice(type_str, tpye_list)
            if t_str not in case_select:
                case_select[t_str] = []
            case_select[t_str].extend(selects)
    return case_select


class AAT_API:

    # resource name
    ue_resource = 'ue'
    epg_resource = 'epg'
    apn_resource = 'apn'
    https_resource = 'httpserver-sut'
    create_https_resource_prefix = 'aat-dpi-http'
    # test step 的 name
    l7_attach_step_name = 'Attach_Request'
    imsi_attach_step_name = '(L) IMSI Attach'
    http_step_name = '(L) Initial HTTP Service'
    https_step_name = 'HTTP(S)_Service'
    ue_detach_step_name = '(L) UE Initiated Detach'
    linux_command_step_name = 'Linux_Command'
    epg_command_step_name = 'Appointed_Command_of_EPG'

    def __init__(self, aat_domain, user_name, password):
        self.aat_domain = aat_domain
        self.user_name = user_name
        self.password = password
        # 取消 https 的告警
        urllib3.disable_warnings()
        # 初始化 api key
        self.api_key = None
        self.get_api_key()
        self.server_id = self.get_epcft_service_id()

    def _common_request(self, method, url, params=None, json=None, headers=None, data=None):
        if not headers:
            headers = {
                "Content_Type": "application/json",
                "apikey": self.api_key,
            }
        else:
            headers["apikey"] = self.api_key
        url = "{}/{}".format(self.aat_domain, url)
        args = {
            "params": params,
            "json": json,
            "headers": headers,
            "data": data,
        }
        response = requests.request(method, url, **args, verify=False, timeout=10)
        status_code = response.status_code
        if status_code >= 400:
            print('url', url)
            print('response.text', response.text)
            raise Exception
        if response.content:
            try:
                return status_code, response.json()
            except JSONDecodeError:
                # print('url', url)
                return status_code, response.text
        return status_code, ''

    def get(self, url, params=None, headers=None):
        return self._common_request("GET", url, params=params, headers=headers)

    def post(self, url, body=None, headers=None, data=None):
        return self._common_request("POST", url, json=body, headers=headers, data=data)

    def delete(self, url, params=None, headers=None):
        return self._common_request("DELETE", url, params=params, headers=headers)

    def get_api_key(self):
        name = self.user_name
        password = self.password
        # 获取 api 调用权限
        url = "/api/aat/apikey?username={}&password={}".format(name, password)
        _, response = self.get(url)
        # print('get_api_key response', response)
        api_key = response['result']['apiKey'][0]
        self.server_id = None
        self.api_key = api_key
        return api_key

    def get_test_services(self):
        # 获取所有 test service 信息
        url = "/api/aat/v1/service"
        _, response = self.get(url)
        # print('get_test_services', response)
        return response['result']

    def get_epcft_service_id(self):
        # 获取 test service 中 epc ft 的信息
        services = self.get_test_services()
        for service_id, value in services.items():
            if 'EPC' in value['serviceName'] and "function test" in value['serviceType']:
                return service_id

    def get_service_test_case(self, service_id):
        url = "/api/aat/v1/{}/test_case".format(service_id)
        _, response = self.get(url)
        return response['result']['data']

    def get_test_case_id(self, service_id, case_name):
        cases = self.get_service_test_case(service_id)
        for case in cases:
            if case_name == case['name']:
                return case['id']
        print("not exist case name:{} in service {}".format(case_name, service_id))
        return

    def get_test_case_by_id(self, service_id, case_id):
        url = "/api/aat/v1/{}/test_case/{}".format(service_id, case_id)
        _, response = self.get(url)
        return response['result']

    def change_test_case_by_id(self, service_id, case_id, change_data):
        url = "/api/aat/v1/{}/test_case/{}".format(service_id, case_id)
        if 'labels' not in change_data:
            change_data['labels'] = []
        self.post(url, change_data)

    def get_resource(self, service_id, type_name):
        # resource_name 就是 resource 的 type
        url = "/frontend/{}/resource/{}".format(service_id, type_name)
        _, response = self.get(url)
        return response['result']

    def get_resource_by_name(self, service_id, type_name, resource_name):
        resources = self.get_resource(service_id, type_name)
        # print('resources', resources)
        for resource in resources:
            name = resource['data']['name']
            if name.upper() == resource_name.upper():
                return resource
        print("not exist resource name:{} in {} type".format(resource_name, type_name))

    def get_resource_by_id(self, service_id, type_name, resource_id):
        url = "/frontend/{}/resource/{}/{}".format(service_id, type_name, resource_id)
        _, response = self.get(url)
        return response['result']

    def change_resource_by_id(self, service_id, type_name, resource_id, change_data):
        url = "/frontend/{}/resource/{}/{}".format(service_id, type_name, resource_id)
        if change_data['description'] == '':
            change_data['description'] = type_name
        _, response = self.post(url, change_data)
        return resource_id

    def create_resource(self, service_id, type_name, data):
        url = "/frontend/{}/resource/{}".format(service_id, type_name)
        _, response = self.post(url, data)
        # resource id
        return response["result"]["id"]

    def create_or_update_resource(self, service_id, type_name, data):
        # 如果存在对应的 resource name，则更新 resource，否则创建 resource
        resource = self.get_resource_by_name(service_id, type_name, data['name'])
        if resource:
            resource_id = self.change_resource_by_id(service_id, type_name, resource['id'], data)
        else:
            resource_id = self.create_resource(service_id, type_name, data)
        return resource_id

    def remove_resource_by_name(self, service_id, type_name, resource_name):
        resource = self.get_resource_by_name(service_id, type_name, resource_name)
        # 存在就删除
        if resource:
            resource_id = resource['id']
            url = "/frontend/{}/resource/{}/{}".format(service_id, type_name, resource_id)
            _, response = self.delete(url)

    def start_test_case_execution(self, service_id, case_id):
        url = "/api/aat/v1/{}/test_execution".format(service_id)
        body = {
            "type": "testcase",
            "id": case_id,
        }
        _, response = self.post(url, body)
        # task id
        return response["result"]["id"]

    def get_test_execution_status_by_id(self, service_id, task_id):
        url = "/api/aat/v1/{}/test_execution?id={}".format(service_id, task_id)
        _, response = self.get(url)
        return response['result']['data']

    def get_test_execution_log_id(self, service_id, task_id, case_id):
        case_done = False
        data = {}
        while not case_done:
            data = self.get_test_execution_status_by_id(service_id, task_id)
            print('case is running, please waiting ...')
            if data.get('endedTime') is None:
                time.sleep(20)
            else:
                case_done = True

        cases = data['testCaseList']
        for case in cases:
            if case_id in [case.get('testcase_id'), case.get('testcaseId')]:
                if not case['logs']:
                    return None
                logs = filter(lambda log: 'Log' in log['name'], case['logs'])
                return list(logs)[0]['id']

    def get_test_case_log(self, service_id, log_id):
        url = "/api/aat/v1/{}/logs/{}?type=text".format(service_id, log_id)
        _, response = self.get(url)
        return response

    def import_test_case(self, service_id, tar_path):
        url = "/frontend/{}/import".format(service_id)
        with open(tar_path, "rb") as tar:
            data_binary = tar.read()
        self.post(url, data=data_binary)

    # def export_test_case(self, service_id):
    #     url = "/frontend/{}/export".format(service_id)


def get_domain_name(url):
    return url.split("://")[1].split("/")[0].split(":")[0]


def url_completion(line):
    url = line['l7']
    if url == "":
        return url

    url = url.replace("//*.", "//www.")
    url = url.replace("//*", "//www.")
    if 'http://' in url:
        # http * 号在尾部则，补充 index
        # http * 号在尾部则删掉 *
        url = url.replace("/*", "/")
        #
        url = url.replace(":*.", ":80")
    else:
        url = url.replace("/*", "/")
        url = url.replace(":*.", ":443")
    # 如果指定了端口后，则在 url 中添加端口号
    if line['l4'] != "" and line['l4'] is not None:
        url = url.replace(".com", ".com:{}".format(line['l4']))
    # print(line['l7'], url)
    return url


def ip_completion(line):
    ip_mask = line.get("l3")
    # print('ip_mask', ip_mask)
    if not ip_mask:
        return None
    if "/" not in ip_mask:
        return ip_mask
    ip, mask = ip_mask.split("/")
    # IPV6，省略全0
    ipv6_colon = ip.count(":")
    ip = ip.replace("::", ":0000" * (8 - ipv6_colon) + ":")
    # 2409:8087::
    if ip[-1] == ":":
        ip += "0000"
    # 唯一地址
    if mask == "32" or mask == "128":
        line["l3"] = ip
    else:
        # 选第一个有效地址作为目标IP
        if ip.count(".") == 0:
            # ipv6
            split_str = ":"
            region_len = 16
        else:
            split_str = "."
            region_len = 8
        ip = ip.split(split_str)
        mask = int(mask)
        mask_range = int(mask / region_len)
        mask_last = int(mask % region_len)
        mask_last = "1" * mask_last + "0" * (region_len - mask_last)
        mask_last = int(mask_last, 2)
        # print("ip[mask_range]", ip[mask_range], mask_range)
        ip[mask_range] = str((int(ip[mask_range]) & mask_last))
        ip[-1] = str(int(ip[-1]) + 1)
        ip = "{}".format(split_str).join(ip)
    return ip


def change_case_epg_and_apn(aat, server_id, steps, config_yaml):
    config_epg_name = config_yaml.get('epg_name')
    epg_with_apn = config_yaml.get('epg_with_apn')
    epg_step = get_test_step_by_name(steps, AAT_API.epg_command_step_name)
    if epg_step:
        step_epg_id = epg_step['parameters']['Target']['value']
        step_epg_name = aat.get_resource_by_id(server_id, aat.epg_resource, step_epg_id)['name']
    else:
        # 如果 case 没有 epg step 则不需要修改 apn 和 epg 的信息
        return steps

    # 如果用户指定了 epg，并且 epg 的名字和 case 模板使用的 epg 名字不同
    if config_epg_name not in ['', None] and config_epg_name != step_epg_name:
        epg = aat.get_resource_by_name(server_id, AAT_API.epg_resource, config_epg_name)
        epg_id = epg['id']
        epg_user = epg['data']['Gom']['User']
        apn_id = aat.get_resource_by_name(server_id, AAT_API.apn_resource, epg_with_apn[config_epg_name])['id']
        # 修改 case 中的 epg 和 apn 的名字
        for step in steps:
            name = step.get('name')
            if step.get('parameters') and 'APN' in step['parameters']:
                step['parameters']['APN']['value'] = apn_id
            if name == aat.l7_attach_step_name:
                step['subParameters']['parameters']['APN']['value'] = apn_id
            if name == aat.epg_command_step_name:
                step['parameters']['Target']['value'] = epg_id
                step['parameters']['User']['value'] = epg_user
    return steps


def change_epg_command_step(aat, server_id, steps):
    imsi = None
    for step in steps:
        name = step.get('name')
        # 获取第一个存在 ue 参数的 step 里面，ue resource 对应的 imsi 的值
        parameters = step.get("parameters")
        if parameters and 'UE' in parameters and not imsi:
            ue_id = step["parameters"]["UE"]["value"]
            ue = aat.get_resource_by_id(server_id, aat.ue_resource, ue_id)
            imsi = ue['IMSI']
        # 修改 epg_command 里面的 command 的值
        if name == AAT_API.epg_command_step_name and imsi:
            command = "epg pgw user-info-sacc identifier-type imsi value {} | nomore".format(imsi)
            step['parameters']['Appointed_Command']['value'] = command
    return steps


def change_https_step_https_resource(aat, server_id, steps, lines, type_str):
    index = 0
    for step in steps:
        name = step.get('name')
        if name == AAT_API.https_step_name:
            #
            resource_name = "{}{}".format(AAT_API.create_https_resource_prefix, index)
            # 默认 80 端口
            port = 80
            if type_str == 'l3':
                # l3 的 case 数据
                node_type = 'HTTPS'
                host_name = ip_completion(lines[index])
            else:
                # l3_l7 或 l7 的 case 数据
                url = url_completion(lines[index])
                node_type = url.split("://")[0].upper()
                host_name = get_domain_name(url)
            # https url 需要将端口改为 443
            if node_type == 'HTTPS':
                port = 443
            # 如果指定端口，则使用指定端口
            if lines[index].get('l4') not in ['', None]:
                port = int(lines[index].get('l4'))
            https_resource_data = {
                '_version': '_removed_intentionally_',
                'HostName': host_name,
                'name': resource_name,
                'Port': port,
                'NodeType': node_type,
                'description': resource_name,
            }
            # 修改 case 的 http resource id
            resource_id = aat.create_or_update_resource(server_id, AAT_API.https_resource, https_resource_data)
            step['destination']['value'] = resource_id
            # l3+l7 以及 l7 的 header 需要带上 Host:域名，这样发包是正确的，才会有业务编码
            if type_str != 'l3':
                step['parameters']['Header'] = {
                    "type": "string",
                    "value": "Host:{}".format(host_name)
                }
            index += 1
    return steps


def remove_https_resource(aat, server_id):
    for index in range(max_modify_step):
        https_resource_name = "{}{}".format(AAT_API.create_https_resource_prefix, index)
        aat.remove_resource_by_name(server_id, AAT_API.https_resource, https_resource_name)


def change_linux_step_command(steps, lines, config_yaml):
    index = 0
    for step in steps:
        name = step.get('name')
        if name == AAT_API.linux_command_step_name:
            ip = ip_completion(lines[index])
            url = url_completion(lines[index])
            # 获取域名
            domain = get_domain_name(url)
            # python3 add_rule_to_etc_hosts.py ip url etc_hosts_path
            # commands = step['parameters']['Commands']['value'].split()
            python = "python3"
            command = "{} {} {} {} {};cat {}|grep {} ".format(
                python, config_yaml['script_path'], ip, domain,
                config_yaml['hosts_path'], config_yaml['hosts_path'], domain
            )
            step['parameters']['Commands']['value'] = command
            index += 1
    return steps


def get_test_step_by_name(steps, step_name):
    for step in steps:
        name = step.get('name')
        if name == step_name:
            return step


def get_step_template(steps, start_step_name, end_step_name=None):
    # 获取 steps 中 start_step_name 到 end_step_name 之前为止的 steps （不包含 end_step_name）
    # 如果没有 end_step_name，则获取到最后的 step
    if end_step_name is None:
        end_step_name = start_step_name

    start = None
    end = len(steps)
    for index, step in enumerate(steps):
        name = step.get('name')
        if name == start_step_name:
            start = index
        elif name == end_step_name and start is not None:
            end = index
            break
    if start is None:
        print("not have step name:{}".format(start_step_name))
        return
    # print("get_step_template", steps)
    return steps[start:end]


def get_step_index(steps, step_name):
    for index, step in enumerate(steps):
        name = step.get('name')
        if name == step_name:
            return index


def delete_step(steps, step_name, delete_len):
    # 删除 steps 中 step name 是 step_name 开始的几个 step
    delete_index = None
    for index, step in enumerate(steps):
        name = step.get('name')
        if name == step_name:
            delete_index = index
    if delete_index is None:
        return
    result = steps[:delete_index]
    result.extend(steps[delete_index + delete_len:])
    return result


def get_step_num(steps, step_name):
    return len([step for step in steps if step.get('name') == step_name])


def verify_case_pass(log, lines):
    # fail: 验证失败, pass: 验证通过
    result = []

    # l3_l7: AAT receive GET <---------- HTTP(S)_Service Response
    # l3: AAT receive TCP_only <---------- HTTP(S)_Service Response
    # l7: AAT <---------- HTTP_Response

    http_respone = "---- HTTP"
    # 鉴权失败
    imsi_fail = "Failed to establish a new connection: [Errno 98] Address already in use"
    # case 在处理数据之前报错
    if log.count("AAT ----------> Attach_Complete") == 0:
        print("case failed befor http step, please check log:\n{}".format(log))
        # 鉴权失败，休眠 5s
        if imsi_fail in log:
            time.sleep(5)
        return result

    def epg_result_check(epg_log, code):
        # if epg_log != "":
        #     return True
        other_code = "3" + code[1:]
        cmnet_flag = "apn-in-use: cmnet" in epg_log
        code_flag = code in epg_log or other_code in epg_log
        if cmnet_flag and code_flag:
            return True
        return False

    datas = log.split(http_respone)[1:]
    for index, data in enumerate(datas):
        code = str(lines[index].get('code'))
        if epg_result_check(data, code):
            result.append('pass')
        else:
            result.append('failed')
    # 最后执行的数据 failed 导致 case 没执行完所有数据
    if len(result) != len(lines):
        result.append('failed')
    return result


def change_case_template_step_num(step_datas, change_step_num, start_step_name, end_step_name=None):
    template_step = get_step_template(step_datas, start_step_name, end_step_name)
    case_template_step_num = get_step_num(step_datas, start_step_name)
    diff_len = change_step_num - case_template_step_num
    while diff_len != 0:
        if diff_len > 0:
            # case https step 少了，需要添加
            if not end_step_name:
                step_datas.extend(template_step)
            else:
                start_index = get_step_index(step_datas, start_step_name)
                start_befor_data = step_datas[:start_index]
                start_after_data = step_datas[start_index:]
                start_befor_data.extend(template_step)
                start_befor_data.extend(start_after_data)
                step_datas = start_befor_data
            diff_len -= 1
        elif diff_len < 0:
            # case https step 多了，需要删除
            step_datas = delete_step(step_datas, start_step_name, len(template_step))
            diff_len += 1
    return step_datas


def change_case_and_execute_and_analyze_log(aat, server_id, case_id, case_data, lines):
    # 修改 case 数据
    aat.change_test_case_by_id(server_id, case_id, case_data)
    # 执行 case
    task_id = aat.start_test_case_execution(server_id, case_id)
    # 获取执行结果
    log_id = aat.get_test_execution_log_id(server_id, task_id, case_id)
    if not log_id:
        return []

    log = aat.get_test_case_log(server_id, log_id)

    # 分析执行结果，将结果更新到 excel 数据
    case_verifys = verify_case_pass(log, lines)

    log_lines = log.split('\n')

    # 在 log 的 epg 命令输出后面，插入提示 "the expected business code is xxx, checking is xxx"
    verify_index = 0
    for index, line in enumerate(log_lines):
        if "epg pgw user-info-sacc identifier-type" in line:
            code = str(lines[verify_index]['code'])
            check = case_verifys[verify_index]
            epg_info = "\nthe expected business code is {} or {}, checking is {}".format(code, '3' + code[1:], check)
            log_lines[index] += epg_info
            verify_index += 1

    log = "\n".join(log_lines)
    print('logID:{}\n'.format(log_id), log)
    return case_verifys


def update_verify_to_line(lines, verify_result, line_index):
    for index, verify in enumerate(verify_result):
        lines[index + line_index]['verify'] = verify
    return lines


def modify_and_execute_case(aat, lines, type_str, config_yaml):
    template_end_step_name = None
    if 'l3' in type_str and 'l7' in type_str:
        # l3+(l4)?+l7+http(s)?
        print('{} {}:\n'.format(type_str, len(lines)), lines)
        case_name = config_yaml['l3_l7']
        template_start_step_name = AAT_API.linux_command_step_name
    elif 'l3' in type_str:
        # l3
        print('{} {}:\n'.format(type_str, len(lines)), lines)
        case_name = config_yaml['l3']
        template_start_step_name = AAT_API.https_step_name
    elif 'l7' in type_str:
        # l7+http(s)?
        print('{} {}:\n'.format(type_str, len(lines)), lines)
        case_name = config_yaml['l7']
        template_start_step_name = AAT_API.https_step_name
        template_end_step_name = AAT_API.ue_detach_step_name

    # 获取 case 的 id
    server_id = aat.server_id
    case_id = aat.get_test_case_id(server_id, case_name)

    # 遍历所有数据，每次最多处理 n 条数据
    index = 0
    # for index in range(0, len(lines), max_modify_step):
    while index < len(lines):
        # 每个 case 间隔 1 秒执行，防止上个 case 失败后，立刻执行下个 case
        time.sleep(1)
        # 获取 case 的 json 数据
        case_data = aat.get_test_case_by_id(server_id, case_id)
        # print('case_data\n', case_data, '\n')
        step_data = case_data['testStepList']
        # 对比 case https step 数量和现在要处理的数据数量，step 数量少了则加 step，多了则减 step
        line_datas = lines[index:index + max_modify_step]
        step_data = change_case_template_step_num(step_data, len(line_datas),
                                                  template_start_step_name, template_end_step_name)

        # 修改 case 的数据
        # 修改 epg 和 apn 在 case 里面的 id
        step_data = change_case_epg_and_apn(aat, server_id, step_data, config_yaml)
        # 修改 epg command step 的内容
        step_data = change_epg_command_step(aat, server_id, step_data)
        if 'l3' in type_str and 'l7' in type_str:
            # 修改 linux command
            step_data = change_linux_step_command(step_data, line_datas, config_yaml)
            # 修改 http resource 的 内容
            step_data = change_https_step_https_resource(aat, server_id, step_data, line_datas, type_str)
        elif 'l3' in type_str:
            step_data = change_https_step_https_resource(aat, server_id, step_data, line_datas, type_str)
        elif 'l7' in type_str:
            step_data = change_https_step_https_resource(aat, server_id, step_data, line_datas, type_str)
        case_data['testStepList'] = step_data

        # 执行 case 并分析结果
        case_verifys = change_case_and_execute_and_analyze_log(aat, server_id, case_id, case_data, line_datas)
        # 如果一条数据都没有执行，则重新跑下 case
        if len(case_verifys) == 0:
            continue
        # 将执行结果更新到 lines 字典中
        update_verify_to_line(lines, case_verifys, index)

        # 下个 case 的数据开始下标，多少条数据执行了，下表就增加多少
        index += len(case_verifys)
        print('Execute {} {} data, {} data remaining'.format(type_str, len(case_verifys), len(lines) - index))

    return lines


def save_data_to_excel(file_path, type_date):
    title = {
        'line_num': '测试数据行数',
        'verify': '验证结果',
        'flag': '变更标志',
        'description': '变更说明',
        'code': '业务编码',
        'name': '业务命名',
        'l3': '三层目的IP地址',
        'protocol': '协议号',
        'l4': '四层目的端口号',
        'l7': '七层URL',
    }

    wb = openpyxl.Workbook()

    sheet = wb.create_sheet(title="输出结果")
    index = 1

    def wirte_one_row(sheet, row_num, datas):
        for index, data in enumerate(datas):
            sheet.cell(row_num, index + 1).value = data

    for type_str, datas in type_date.items():
        if index == 1:
            # 写表头
            keys = list(datas[0].keys())
            # 添加验证结果
            if "verify" not in keys:
                keys.append("verify")
            # 英文转中文
            for key_index, key in enumerate(keys):
                if key in title:
                    keys[key_index] = title[key]
            wirte_one_row(sheet, index, keys)
            index += 1
        # 写数据
        for data in datas:
            if "verify" not in data:
                # 没执行过的数据，不写入 excel
                continue
            wirte_one_row(sheet, index, data.values())
            index += 1
    wb.save(file_path)


if __name__ == "__main__":
    try:
        # aat_dpi input_path output_path
        if len(sys.argv) < 3:
            print("argv:{}. less then 3 argument. "
                  "example: ./aat_dpi ./aat_dpi.xlsx ./date_dpi_output.xlsx".format(sys.argv))
            sys.exit()
        input_path = sys.argv[1]
        output_path = sys.argv[2]

        # 检查是否提前建立好 dns 文件
        # DNS_TRANSLATION_FILE = './aat_dpi'
        DNS_TRANSLATION_FILE = '/aat_dpi/hosts'
        ls_dns = os.popen('ls {}'.format(DNS_TRANSLATION_FILE)).read()
        dns_exist = (ls_dns == "{}\n".format(DNS_TRANSLATION_FILE))
        if not dns_exist:
            print('please check {} is exist, or create a empty file'.format(DNS_TRANSLATION_FILE))
            sys.exit()

        # 读取配置文件
        with open('./config.yaml', 'r') as config:
            config_yaml = load(config, Loader=FullLoader)
        config_yaml['hosts_path'] = DNS_TRANSLATION_FILE
        config_yaml['input_path'] = input_path
        config_yaml['output_path'] = output_path
        print('config_yaml', config_yaml)

        # 读取 excel
        type_date = read_dpi_config(config_yaml['input_path'], config_yaml['date_select_max_num'])
        # print(type_date)

        # 打印选中数据的数量
        for key in type_date:
            print(key, len(type_date[key]))

        # 初始化 aat api
        aat = AAT_API(config_yaml['aat_domain'], config_yaml['user_name'], config_yaml['password'])

        # 根据 excel 数据修改 case，执行并获取 case 执行结果
        for type_str, lines in type_date.items():
            modify_and_execute_case(aat, lines, type_str, config_yaml)

    finally:
        # # 删除可能创建的 https resource
        # remove_https_resource(aat, aat.server_id)

        # 将执行结果输出到 excel
        save_data_to_excel(config_yaml['output_path'], type_date)

        #
        input("Enter any key to finish the program")
